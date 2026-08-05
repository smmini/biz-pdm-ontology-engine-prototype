import os
import openpyxl

def raw_preview(file_path: str) -> dict:
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".csv", ".txt", ".tsv"):
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            lines = [f.readline() for _ in range(20)]
        return {"file_type": "text", "extension": ext, "raw_lines": lines}

    elif ext in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        preview = {}
        for name in wb.sheetnames:
            rows = []
            for i, row in enumerate(wb[name].iter_rows(values_only=True)):
                if i >= 15:
                    break
                rows.append(list(row))
            preview[name] = rows
        return {"file_type": "excel", "sheet_names": wb.sheetnames, "raw_preview": preview}

    elif ext == ".json":
        with open(file_path, "r", encoding="utf-8") as f:
            return {"file_type": "json", "raw_text": f.read(4000)}

    else:
        return {"file_type": "unknown", "extension": ext}
